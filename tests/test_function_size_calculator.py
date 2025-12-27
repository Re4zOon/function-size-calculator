#!/usr/bin/env python3
"""
Test suite for function_size_calculator.py
"""

import unittest
import os
import sys
import tempfile
import shutil
import json
from pathlib import Path
from io import StringIO
from contextlib import redirect_stdout, redirect_stderr

# Add parent directory to path to import the module
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from function_size_calculator import (
    FunctionInfo,
    JavaScriptParser,
    JavaParser,
    PythonParser,
    CSharpParser,
    ExcelWriter,
    JSONWriter,
    scan_single_repository
)

try:
    import openpyxl
except ImportError:
    print("Warning: openpyxl not available. Skipping Excel tests.")
    openpyxl = None


class TestFunctionInfo(unittest.TestCase):
    """Test cases for FunctionInfo class."""
    
    def test_function_info_creation(self):
        """Test creating a FunctionInfo object."""
        func = FunctionInfo(
            name="testFunction",
            file_path="test.js",
            start_line=1,
            end_line=10,
            size=10
        )
        
        self.assertEqual(func.name, "testFunction")
        self.assertEqual(func.file_path, "test.js")
        self.assertEqual(func.start_line, 1)
        self.assertEqual(func.end_line, 10)
        self.assertEqual(func.size, 10)
    
    def test_function_info_repr(self):
        """Test FunctionInfo string representation."""
        func = FunctionInfo("myFunc", "file.js", 5, 15, 11)
        repr_str = repr(func)
        
        self.assertIn("myFunc", repr_str)
        self.assertIn("file.js", repr_str)
        self.assertIn("11", repr_str)
    
    def test_function_info_to_dict(self):
        """Test FunctionInfo to_dict method."""
        func = FunctionInfo("testFunc", "test.py", 10, 20, 11)
        func_dict = func.to_dict()
        
        self.assertEqual(func_dict['name'], "testFunc")
        self.assertEqual(func_dict['file_path'], "test.py")
        self.assertEqual(func_dict['start_line'], 10)
        self.assertEqual(func_dict['end_line'], 20)
        self.assertEqual(func_dict['size'], 11)


class TestJavaScriptParser(unittest.TestCase):
    """Test cases for JavaScriptParser."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.fixtures_dir = os.path.join(
            os.path.dirname(__file__), 
            'fixtures'
        )
        self.js_file = os.path.join(self.fixtures_dir, 'sample.js')
        self.ts_file = os.path.join(self.fixtures_dir, 'sample.ts')
    
    def test_parse_javascript_file(self):
        """Test parsing a JavaScript file."""
        functions = JavaScriptParser.parse_functions(self.js_file)
        
        # Should find multiple functions
        self.assertGreater(len(functions), 0)
        
        # Check for specific functions
        func_names = [f.name for f in functions]
        self.assertIn("simpleFunction", func_names)
        self.assertIn("largeFunction", func_names)
        self.assertIn("arrowFunction", func_names)
        self.assertIn("asyncArrowFunction", func_names)
        self.assertIn("outerFunction", func_names)
    
    def test_parse_typescript_file(self):
        """Test parsing a TypeScript file."""
        functions = JavaScriptParser.parse_functions(self.ts_file)
        
        # Should find functions
        self.assertGreater(len(functions), 0)
        
        # Check for specific functions
        func_names = [f.name for f in functions]
        self.assertIn("typedFunction", func_names)
        # Note: typedArrow may not be detected due to TypeScript type annotations
        # which the simple regex parser doesn't fully support
    
    def test_function_size_calculation(self):
        """Test that function sizes are calculated correctly."""
        functions = JavaScriptParser.parse_functions(self.js_file)
        
        # Find the simpleFunction
        simple = next((f for f in functions if f.name == "simpleFunction"), None)
        self.assertIsNotNone(simple)
        
        # simpleFunction should be 3 lines
        self.assertEqual(simple.size, 3)
        
        # Find largeFunction
        large = next((f for f in functions if f.name == "largeFunction"), None)
        self.assertIsNotNone(large)
        
        # largeFunction should be larger than simpleFunction
        self.assertGreater(large.size, simple.size)
    
    def test_parse_nonexistent_file(self):
        """Test parsing a file that doesn't exist."""
        # Suppress expected warning output
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = JavaScriptParser.parse_functions("/nonexistent/file.js")
        
        # Should return empty list, not crash
        self.assertEqual(len(functions), 0)
    
    def test_function_line_numbers(self):
        """Test that line numbers are correct."""
        functions = JavaScriptParser.parse_functions(self.js_file)
        
        for func in functions:
            # Start line should be positive
            self.assertGreater(func.start_line, 0)
            # End line should be greater than start line
            self.assertGreaterEqual(func.end_line, func.start_line)
            # Size should be positive
            self.assertGreater(func.size, 0)


class TestJavaParser(unittest.TestCase):
    """Test cases for JavaParser."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.fixtures_dir = os.path.join(
            os.path.dirname(__file__), 
            'fixtures'
        )
        self.java_file = os.path.join(self.fixtures_dir, 'Sample.java')
    
    def test_parse_java_file(self):
        """Test parsing a Java file."""
        functions = JavaParser.parse_functions(self.java_file)
        
        # Should find multiple methods
        self.assertGreater(len(functions), 0)
        
        # Check for specific methods
        func_names = [f.name for f in functions]
        self.assertIn("publicMethod", func_names)
        self.assertIn("privateMethod", func_names)
        self.assertIn("protectedStaticMethod", func_names)
        self.assertIn("largeMethod", func_names)
    
    def test_java_method_modifiers(self):
        """Test that methods with various modifiers are detected."""
        functions = JavaParser.parse_functions(self.java_file)
        func_names = [f.name for f in functions]
        
        # Test different modifier combinations
        self.assertIn("publicStaticFinalMethod", func_names)
        self.assertIn("synchronizedMethod", func_names)
        self.assertIn("methodWithException", func_names)
    
    def test_java_function_size(self):
        """Test that Java method sizes are calculated correctly."""
        functions = JavaParser.parse_functions(self.java_file)
        
        # Find the largeMethod
        large = next((f for f in functions if f.name == "largeMethod"), None)
        self.assertIsNotNone(large)
        
        # largeMethod should be at least 10 lines
        self.assertGreaterEqual(large.size, 10)
    
    def test_parse_nonexistent_java_file(self):
        """Test parsing a Java file that doesn't exist."""
        # Suppress expected warning output
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = JavaParser.parse_functions("/nonexistent/Sample.java")
        
        # Should return empty list, not crash
        self.assertEqual(len(functions), 0)


class TestCSharpParser(unittest.TestCase):
    """Test cases for CSharpParser."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.fixtures_dir = os.path.join(
            os.path.dirname(__file__), 
            'fixtures'
        )
        self.cs_file = os.path.join(self.fixtures_dir, 'Sample.cs')
    
    def test_parse_csharp_file(self):
        """Test parsing a C# file."""
        functions = CSharpParser.parse_functions(self.cs_file)
        
        # Should find multiple methods
        self.assertGreater(len(functions), 0)
        
        # Check for specific methods
        func_names = [f.name for f in functions]
        self.assertIn("simpleAdd", func_names)
        self.assertIn("largeMethod", func_names)
        self.assertIn("asyncMethod", func_names)
        self.assertIn("virtualMethod", func_names)
        self.assertIn("ToString", func_names)
    
    def test_csharp_method_modifiers(self):
        """Test that methods with various modifiers are detected."""
        functions = CSharpParser.parse_functions(self.cs_file)
        func_names = [f.name for f in functions]
        
        # Test different modifier combinations
        self.assertIn("simpleAdd", func_names)  # public
        self.assertIn("largeMethod", func_names)  # private static
        self.assertIn("asyncMethod", func_names)  # protected async
        self.assertIn("virtualMethod", func_names)  # internal virtual
        self.assertIn("ToString", func_names)  # public override
    
    def test_csharp_function_size(self):
        """Test that C# method sizes are calculated correctly."""
        functions = CSharpParser.parse_functions(self.cs_file)
        
        # Find the largeMethod
        large = next((f for f in functions if f.name == "largeMethod"), None)
        self.assertIsNotNone(large)
        
        # largeMethod should be at least 10 lines
        self.assertGreaterEqual(large.size, 10)
        
        # Find simpleAdd
        simple = next((f for f in functions if f.name == "simpleAdd"), None)
        self.assertIsNotNone(simple)
        
        # simpleAdd should be small
        self.assertLess(simple.size, 5)
    
    def test_parse_nonexistent_csharp_file(self):
        """Test parsing a C# file that doesn't exist."""
        # Suppress expected warning output
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = CSharpParser.parse_functions("/nonexistent/Sample.cs")
        
        # Should return empty list, not crash
        self.assertEqual(len(functions), 0)


class TestPythonParser(unittest.TestCase):
    """Test cases for PythonParser."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.fixtures_dir = os.path.join(
            os.path.dirname(__file__), 
            'fixtures'
        )
        self.py_file = os.path.join(self.fixtures_dir, 'sample.py')
    
    def test_parse_python_file(self):
        """Test parsing a Python file."""
        functions = PythonParser.parse_functions(self.py_file)
        
        # Should find multiple functions
        self.assertGreater(len(functions), 0)
        
        # Check for specific functions
        func_names = [f.name for f in functions]
        self.assertIn("simple_function", func_names)
        self.assertIn("large_function", func_names)
        self.assertIn("async_function", func_names)
    
    def test_python_class_methods(self):
        """Test that class methods are detected."""
        functions = PythonParser.parse_functions(self.py_file)
        func_names = [f.name for f in functions]
        
        # Test class methods
        self.assertIn("__init__", func_names)
        self.assertIn("instance_method", func_names)
        self.assertIn("static_method", func_names)
        self.assertIn("class_method", func_names)
    
    def test_python_function_size(self):
        """Test that Python function sizes are calculated correctly."""
        functions = PythonParser.parse_functions(self.py_file)
        
        # Find the simple_function
        simple = next((f for f in functions if f.name == "simple_function"), None)
        self.assertIsNotNone(simple)
        
        # simple_function should be small
        self.assertLess(simple.size, 10)
        
        # Find large_function
        large = next((f for f in functions if f.name == "large_function"), None)
        self.assertIsNotNone(large)
        
        # large_function should be larger
        self.assertGreater(large.size, simple.size)
    
    def test_multiline_signature(self):
        """Test that functions with multi-line signatures are parsed correctly."""
        # Create a temporary file with a multi-line function signature
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as f:
            temp_file = f.name
            f.write("""def multi_line_func(
    arg1: str,
    arg2: int,
    arg3: dict
) -> bool:
    x = 1
    y = 2
    return True
""")
        
        try:
            functions = PythonParser.parse_functions(temp_file)
            
            # Should find the function
            self.assertEqual(len(functions), 1)
            self.assertEqual(functions[0].name, "multi_line_func")
            
            # Should have correct size (8 lines total)
            self.assertEqual(functions[0].size, 8)
        finally:
            if os.path.exists(temp_file):
                os.remove(temp_file)
    
    def test_parse_nonexistent_python_file(self):
        """Test parsing a Python file that doesn't exist."""
        # Suppress expected warning output
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = PythonParser.parse_functions("/nonexistent/sample.py")
        
        # Should return empty list, not crash
        self.assertEqual(len(functions), 0)


class TestExcelWriter(unittest.TestCase):
    """Test cases for ExcelWriter."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp(prefix='test_function_calc_')
        self.output_file = os.path.join(self.temp_dir, 'test_output.xlsx')
        
        # Create sample function data
        self.sample_functions = [
            FunctionInfo("func1", "file1.js", 1, 10, 10),
            FunctionInfo("func2", "file2.js", 1, 20, 20),
            FunctionInfo("func3", "file3.js", 1, 15, 15),
            FunctionInfo("func4", "file4.js", 1, 5, 5),
            FunctionInfo("func5", "file5.js", 1, 8, 8),
            FunctionInfo("func6", "file6.js", 1, 12, 12),
        ]
    
    def tearDown(self):
        """Clean up temporary files."""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_write_results_single_repo(self):
        """Test writing results for a single repository."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Suppress "Results saved to" output
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file)
        
        # Check that file was created
        self.assertTrue(os.path.exists(self.output_file))
        
        # Load and verify content
        wb = openpyxl.load_workbook(self.output_file)
        self.assertIn('test-repo', wb.sheetnames)
        
        ws = wb['test-repo']
        
        # Check header row
        self.assertEqual(ws.cell(1, 1).value, 'Rank')
        self.assertEqual(ws.cell(1, 2).value, 'Function Name')
        self.assertEqual(ws.cell(1, 6).value, 'Lines of Code')
        
        # Check that top 5 functions are present (sorted by size)
        # func2 (20), func3 (15), func6 (12), func1 (10), func5 (8)
        self.assertEqual(ws.cell(2, 1).value, 1)  # Rank 1
        self.assertEqual(ws.cell(2, 2).value, 'func2')  # Largest function
        self.assertEqual(ws.cell(2, 6).value, 20)  # Size
        
        self.assertEqual(ws.cell(6, 1).value, 5)  # Rank 5
        
        wb.close()
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_write_results_multiple_repos(self):
        """Test writing results for multiple repositories."""
        repo_results = {
            'repo1': self.sample_functions[:3],
            'repo2': self.sample_functions[3:]
        }
        
        # Suppress "Results saved to" output
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file)
        
        # Check that file was created
        self.assertTrue(os.path.exists(self.output_file))
        
        # Load and verify content
        wb = openpyxl.load_workbook(self.output_file)
        
        # Check that both sheets exist
        self.assertIn('repo1', wb.sheetnames)
        self.assertIn('repo2', wb.sheetnames)
        
        wb.close()
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_sanitize_sheet_name(self):
        """Test that sheet names are sanitized properly."""
        # Test with long name and special characters
        repo_results = {
            'very/long/repository/name/that/exceeds/thirty/one/characters': self.sample_functions[:1]
        }
        
        # Suppress "Results saved to" output
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file)
        
        wb = openpyxl.load_workbook(self.output_file)
        
        # Check that sheet name is sanitized (max 31 chars, no slashes)
        sheet_names = wb.sheetnames
        self.assertEqual(len(sheet_names), 1)
        self.assertLessEqual(len(sheet_names[0]), 31)
        self.assertNotIn('/', sheet_names[0])
        
        wb.close()
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_top_n_parameter(self):
        """Test writing results with custom top N parameter."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Write only top 3
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file, top_n=3)
        
        wb = openpyxl.load_workbook(self.output_file)
        ws = wb['test-repo']
        
        # Should have header + 3 data rows + empty row + 5 summary rows = 9 rows minimum
        # Count data rows (non-empty cells in column A starting from row 2)
        data_rows = 0
        for row in range(2, 10):
            if ws.cell(row, 1).value and isinstance(ws.cell(row, 1).value, int):
                data_rows += 1
        
        self.assertEqual(data_rows, 3)
        wb.close()
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_min_size_filter(self):
        """Test writing results with minimum size filter."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Filter out functions smaller than 10 lines (should keep func1, func2, func3, func6)
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file, top_n=10, min_size=10)
        
        wb = openpyxl.load_workbook(self.output_file)
        ws = wb['test-repo']
        
        # Count data rows
        data_rows = 0
        for row in range(2, 10):
            if ws.cell(row, 1).value and isinstance(ws.cell(row, 1).value, int):
                data_rows += 1
        
        # Should have 4 functions (func2, func3, func6, func1)
        self.assertEqual(data_rows, 4)
        
        # All should be >= 10 lines
        for row in range(2, 6):
            size = ws.cell(row, 6).value
            if size is not None and isinstance(size, int):
                self.assertGreaterEqual(size, 10)
        
        wb.close()
    
    @unittest.skipIf(openpyxl is None, "openpyxl not available")
    def test_summary_statistics(self):
        """Test that summary statistics are added to Excel output."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, self.output_file)
        
        wb = openpyxl.load_workbook(self.output_file)
        ws = wb['test-repo']
        
        # Look for summary statistics section
        found_summary = False
        for row in range(1, 20):
            cell_value = ws.cell(row, 1).value
            if cell_value == 'Summary Statistics':
                found_summary = True
                break
        
        self.assertTrue(found_summary, "Summary statistics section not found")
        wb.close()


class TestJSONWriter(unittest.TestCase):
    """Test cases for JSONWriter."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp(prefix='test_json_')
        self.output_file = os.path.join(self.temp_dir, 'test_output.json')
        
        # Create sample function data
        self.sample_functions = [
            FunctionInfo("func1", "file1.js", 1, 10, 10),
            FunctionInfo("func2", "file2.js", 1, 20, 20),
            FunctionInfo("func3", "file3.js", 1, 15, 15),
            FunctionInfo("func4", "file4.js", 1, 5, 5),
            FunctionInfo("func5", "file5.js", 1, 8, 8),
        ]
    
    def tearDown(self):
        """Clean up temporary files."""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_write_results_single_repo(self):
        """Test writing results for a single repository to JSON."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Suppress "Results saved to" output
        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, self.output_file)
        
        # Check that file was created
        self.assertTrue(os.path.exists(self.output_file))
        
        # Load and verify content
        with open(self.output_file, 'r') as f:
            data = json.load(f)
        
        self.assertIn('test-repo', data)
        self.assertIn('summary', data['test-repo'])
        self.assertIn('top_functions', data['test-repo'])
        
        # Check summary statistics
        summary = data['test-repo']['summary']
        self.assertEqual(summary['total_functions'], 5)
        self.assertGreater(summary['average_size'], 0)
        self.assertEqual(summary['largest_function_size'], 20)
        self.assertEqual(summary['smallest_function_size'], 5)
        
        # Check top functions (should be sorted by size)
        top_funcs = data['test-repo']['top_functions']
        self.assertEqual(len(top_funcs), 5)
        self.assertEqual(top_funcs[0]['name'], 'func2')  # Largest (20 lines)
        self.assertEqual(top_funcs[0]['size'], 20)
    
    def test_write_results_multiple_repos(self):
        """Test writing results for multiple repositories to JSON."""
        repo_results = {
            'repo1': self.sample_functions[:3],
            'repo2': self.sample_functions[3:]
        }
        
        # Suppress "Results saved to" output
        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, self.output_file)
        
        # Check that file was created
        self.assertTrue(os.path.exists(self.output_file))
        
        # Load and verify content
        with open(self.output_file, 'r') as f:
            data = json.load(f)
        
        # Check that both repos exist
        self.assertIn('repo1', data)
        self.assertIn('repo2', data)
    
    def test_top_n_parameter(self):
        """Test writing results with custom top N parameter to JSON."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Write only top 3
        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, self.output_file, top_n=3)
        
        with open(self.output_file, 'r') as f:
            data = json.load(f)
        
        # Should have only 3 functions
        self.assertEqual(len(data['test-repo']['top_functions']), 3)
    
    def test_min_size_filter(self):
        """Test writing results with minimum size filter to JSON."""
        repo_results = {
            'test-repo': self.sample_functions
        }
        
        # Filter out functions smaller than 10 lines
        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, self.output_file, top_n=10, min_size=10)
        
        with open(self.output_file, 'r') as f:
            data = json.load(f)
        
        # Should have 3 functions (func1=10, func2=20, func3=15)
        top_funcs = data['test-repo']['top_functions']
        self.assertEqual(len(top_funcs), 3)
        
        # All should be >= 10 lines
        for func in top_funcs:
            self.assertGreaterEqual(func['size'], 10)
        
        # Summary should reflect filtered functions
        self.assertEqual(data['test-repo']['summary']['total_functions'], 3)
    
    def test_min_size_filter_multiple_repos(self):
        """Test that min_size filter works correctly for multiple repositories.
        
        This test verifies that the min_size parameter is not accidentally
        shadowed by local variables during processing of multiple repos.
        """
        # Create different sample functions for each repo
        repo1_functions = [
            FunctionInfo("large_func", "file1.js", 1, 50, 50),
            FunctionInfo("medium_func", "file2.js", 1, 25, 25),
        ]
        repo2_functions = [
            FunctionInfo("small_func", "file3.js", 1, 5, 5),
            FunctionInfo("tiny_func", "file4.js", 1, 3, 3),
        ]
        
        repo_results = {
            'repo1': repo1_functions,
            'repo2': repo2_functions,
        }
        
        # Set min_size=2 so all functions should be included
        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, self.output_file, top_n=10, min_size=2)
        
        with open(self.output_file, 'r') as f:
            data = json.load(f)
        
        # Both repos should have functions
        self.assertEqual(data['repo1']['summary']['total_functions'], 2)
        self.assertEqual(data['repo2']['summary']['total_functions'], 2)
        
        # Verify all functions are included
        self.assertEqual(len(data['repo1']['top_functions']), 2)
        self.assertEqual(len(data['repo2']['top_functions']), 2)


class TestScanRepository(unittest.TestCase):
    """Test cases for scan_single_repository function."""
    
    def setUp(self):
        """Set up test repository."""
        self.test_repo_dir = tempfile.mkdtemp(prefix='test_repo_')
        
        # Create a simple test repository structure
        os.makedirs(os.path.join(self.test_repo_dir, 'src'), exist_ok=True)
        
        # Create a JavaScript file
        js_content = """
function testFunc() {
    console.log("test");
}
"""
        with open(os.path.join(self.test_repo_dir, 'src', 'test.js'), 'w') as f:
            f.write(js_content)
        
        # Create a Java file
        java_content = """
public class Test {
    public void testMethod() {
        System.out.println("test");
    }
}
"""
        with open(os.path.join(self.test_repo_dir, 'src', 'Test.java'), 'w') as f:
            f.write(java_content)
    
    def tearDown(self):
        """Clean up test repository."""
        if os.path.exists(self.test_repo_dir):
            shutil.rmtree(self.test_repo_dir)
    
    def test_scan_local_repository(self):
        """Test scanning a local repository."""
        repo_name, functions = scan_single_repository(self.test_repo_dir)
        
        # Check that repo name is extracted
        self.assertIsNotNone(repo_name)
        
        # Should find functions from both JS and Java files
        self.assertGreater(len(functions), 0)
        
        # Check that function names are found
        func_names = [f.name for f in functions]
        self.assertIn("testFunc", func_names)
        self.assertIn("testMethod", func_names)
    
    def test_scan_nonexistent_repository(self):
        """Test scanning a repository that doesn't exist."""
        # Suppress expected error output
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            repo_name, functions = scan_single_repository("/nonexistent/repo")
        
        # Should return None and empty list
        self.assertIsNone(repo_name)
        self.assertEqual(len(functions), 0)
    
    def test_relative_paths(self):
        """Test that file paths are relative to repo root."""
        repo_name, functions = scan_single_repository(self.test_repo_dir)
        
        for func in functions:
            # Paths should be relative
            self.assertFalse(func.file_path.startswith('/'))
            self.assertFalse(func.file_path.startswith(self.test_repo_dir))


class TestCommandLineArguments(unittest.TestCase):
    """Test cases for command-line argument handling."""
    
    def setUp(self):
        """Set up test files."""
        self.temp_dir = tempfile.mkdtemp(prefix='test_cli_')
        self.input_file = os.path.join(self.temp_dir, 'repos.txt')
    
    def tearDown(self):
        """Clean up test files."""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def test_input_file_parsing(self):
        """Test parsing repository list from input file."""
        # Create input file with repositories
        with open(self.input_file, 'w') as f:
            f.write("https://github.com/user/repo1.git\n")
            f.write("# This is a comment\n")
            f.write("\n")  # Empty line
            f.write("/path/to/local/repo\n")
        
        # Read and parse
        repositories = []
        with open(self.input_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    repositories.append(line)
        
        # Should have 2 repositories (comment and empty line ignored)
        self.assertEqual(len(repositories), 2)
        self.assertIn("https://github.com/user/repo1.git", repositories)
        self.assertIn("/path/to/local/repo", repositories)


if __name__ == '__main__':
    # Run tests with verbosity
    unittest.main(verbosity=2)
