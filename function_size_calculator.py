#!/usr/bin/env python3
"""
Function Size Calculator
Scans git repositories to find the largest functions in Node.js, Java, and Python codebases.
Outputs results to an Excel (XLSX) file with each repository on a separate tab.
"""

import argparse
import os
import re
import sys
import tempfile
import shutil
import json
from pathlib import Path
from typing import List, Dict, Tuple
import subprocess
from concurrent.futures import ProcessPoolExecutor, as_completed

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


def print_progress_bar(current: int, total: int, prefix: str = '', suffix: str = '', length: int = 50):
    """
    Print a progress bar to the console.
    
    Args:
        current: Current progress value
        total: Total value for 100% completion
        prefix: Text to display before the progress bar
        suffix: Text to display after the progress bar
        length: Character length of the progress bar
    """
    if total == 0:
        return
    
    percent = 100 * (current / float(total))
    filled_length = int(length * current // total)
    bar = '█' * filled_length + '-' * (length - filled_length)
    
    # Use \r to return to the start of the line and overwrite
    print(f'\r{prefix} |{bar}| {percent:.1f}% {suffix}', end='', flush=True)
    
    # Print newline when complete
    if current == total:
        print()


class FunctionInfo:
    """
    Represents information about a function or method.
    
    Attributes:
        name: The name of the function/method
        file_path: Path to the file containing the function
        start_line: Line number where the function starts (1-indexed)
        end_line: Line number where the function ends (1-indexed)
        size: Total number of lines in the function
    """
    
    def __init__(self, name: str, file_path: str, start_line: int, end_line: int, size: int):
        self.name = name
        self.file_path = file_path
        self.start_line = start_line
        self.end_line = end_line
        self.size = size
    
    def __repr__(self):
        return f"FunctionInfo({self.name}, {self.file_path}, lines={self.size})"
    
    def to_dict(self) -> Dict:
        """Convert FunctionInfo to dictionary for JSON serialization."""
        return {
            'name': self.name,
            'file_path': self.file_path,
            'start_line': self.start_line,
            'end_line': self.end_line,
            'size': self.size
        }


class JavaScriptParser:
    """Parser for JavaScript/TypeScript functions."""
    
    # Compile regex patterns once for better performance
    PATTERNS = [
        # function declaration: function name() {}
        (re.compile(r'^\s*function\s+(\w+)\s*\('), 'function'),
        # arrow function: const name = () => {}
        (re.compile(r'^\s*(?:const|let|var)\s+(\w+)\s*=\s*(?:async\s*)?\([^)]*\)\s*=>'), 'arrow'),
        # method: name() {}
        (re.compile(r'^\s*(?:async\s+)?(\w+)\s*\([^)]*\)\s*\{'), 'method'),
        # class method: async name() {}
        (re.compile(r'^\s*(?:public|private|protected|static)?\s*(?:async\s+)?(\w+)\s*\([^)]*\)\s*\{'), 'class_method'),
    ]
    
    @staticmethod
    def parse_functions(file_path: str) -> List[FunctionInfo]:
        """
        Parse JavaScript/TypeScript file to extract functions.
        
        Uses a streaming approach to handle very large files efficiently
        without loading the entire file into memory.
        
        Supports various function patterns including:
        - Function declarations: function name() {}
        - Arrow functions: const name = () => {}
        - Methods: name() {}
        - Class methods with modifiers: async name() {}
        
        Args:
            file_path: Path to the JavaScript/TypeScript file
            
        Returns:
            List of FunctionInfo objects for all detected functions
        """
        functions = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                line_num = 0
                current_function = None  # (name, start_line, brace_count)
                
                for line in f:  # Stream file line by line
                    line_num += 1
                    
                    # If we're currently tracking a function, update brace count
                    if current_function:
                        func_name, start_line, brace_count = current_function
                        brace_count += line.count('{') - line.count('}')
                        
                        if brace_count == 0:
                            # Function ended
                            size = line_num - start_line + 1
                            functions.append(FunctionInfo(
                                name=func_name,
                                file_path=file_path,
                                start_line=start_line,
                                end_line=line_num,
                                size=size
                            ))
                            current_function = None
                        else:
                            current_function = (func_name, start_line, brace_count)
                    else:
                        # Look for new function declarations
                        for pattern, func_type in JavaScriptParser.PATTERNS:
                            match = pattern.search(line)
                            if match:
                                func_name = match.group(1)
                                brace_count = line.count('{') - line.count('}')
                                
                                if brace_count == 0:
                                    # Single-line function (rare but possible)
                                    functions.append(FunctionInfo(
                                        name=func_name,
                                        file_path=file_path,
                                        start_line=line_num,
                                        end_line=line_num,
                                        size=1
                                    ))
                                elif brace_count > 0:
                                    # Multi-line function - start tracking
                                    current_function = (func_name, line_num, brace_count)
                                break
                                
        except Exception as e:
            print(f"Warning: Could not read {file_path}: {e}")
        
        return functions


class PythonParser:
    """Parser for Python functions."""
    
    # Compile regex pattern once for better performance
    FUNC_PATTERN = re.compile(r'^\s*(?:async\s+)?def\s+(\w+)\s*\(')
    
    @staticmethod
    def parse_functions(file_path: str) -> List[FunctionInfo]:
        """
        Parse Python file to extract functions.
        
        Note: Due to Python's indentation-based syntax requiring lookahead
        to determine function boundaries, this parser reads all lines into
        memory. For extremely large Python files (100MB+), this may cause
        memory pressure. However, such large single-file Python modules are
        rare in practice.
        
        Args:
            file_path: Path to the Python file
            
        Returns:
            List of FunctionInfo objects for all detected functions
        """
        functions = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = []
                line_num = 0
                
                # Read file line by line to avoid loading entire file into memory
                for line in f:
                    lines.append(line.rstrip('\n\r'))
                    line_num += 1
                
                # Now process the lines (this is necessary for Python due to 
                # indentation-based scoping which requires look-ahead)
                i = 0
                while i < len(lines):
                    line = lines[i]
                    
                    match = PythonParser.FUNC_PATTERN.search(line)
                    if match:
                        func_name = match.group(1)
                        start_line = i + 1
                        
                        # Get the base indentation level of the function definition
                        base_indent = len(line) - len(line.lstrip())
                        
                        # Find where the function signature ends (look for ':')
                        # Handle multi-line signatures by checking for unmatched parentheses
                        j = i
                        paren_count = 0
                        while j < len(lines):
                            paren_count += lines[j].count('(') - lines[j].count(')')
                            if ':' in lines[j] and paren_count == 0:
                                break
                            j += 1
                        
                        # Now find where the function body ends
                        j += 1
                        end_line = i  # Initialize to function start (0-indexed)
                        
                        while j < len(lines):
                            current_line = lines[j]
                            
                            # Skip blank lines and comments
                            if current_line.strip() == '' or current_line.strip().startswith('#'):
                                j += 1
                                continue
                            
                            # Check indentation
                            current_indent = len(current_line) - len(current_line.lstrip())
                            
                            # If we're back to the base level or lower, function has ended
                            if current_indent <= base_indent:
                                break
                            
                            end_line = j
                            j += 1
                        
                        if end_line >= i:
                            size = end_line - i + 1
                            functions.append(FunctionInfo(
                                name=func_name,
                                file_path=file_path,
                                start_line=i + 1,  # Convert to 1-indexed
                                end_line=end_line + 1,  # Convert to 1-indexed
                                size=size
                            ))
                    i += 1
                                
        except Exception as e:
            print(f"Warning: Could not read {file_path}: {e}")
        
        return functions


class JavaParser:
    """Parser for Java functions/methods."""
    
    # Compile regex pattern once for better performance
    METHOD_PATTERN = re.compile(
        r'^\s*(?:public|private|protected)?\s*(?:static)?\s*(?:final)?\s*(?:synchronized)?\s*'
        r'[\w<>\[\]]+\s+(\w+)\s*\([^)]*\)\s*(?:throws\s+[\w\s,]+)?\s*\{'
    )
    
    @staticmethod
    def parse_functions(file_path: str) -> List[FunctionInfo]:
        """
        Parse Java file to extract methods.
        
        Uses a streaming approach to handle very large files efficiently
        without loading the entire file into memory.
        
        Supports methods with various modifiers including:
        - Access modifiers: public, private, protected
        - Other modifiers: static, final, synchronized
        - Generic return types
        - Throws clauses
        
        Args:
            file_path: Path to the Java file
            
        Returns:
            List of FunctionInfo objects for all detected methods
        """
        functions = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                line_num = 0
                current_method = None  # (name, start_line, brace_count)
                
                for line in f:  # Stream file line by line
                    line_num += 1
                    
                    # If we're currently tracking a method, update brace count
                    if current_method:
                        method_name, start_line, brace_count = current_method
                        brace_count += line.count('{') - line.count('}')
                        
                        if brace_count == 0:
                            # Method ended
                            size = line_num - start_line + 1
                            functions.append(FunctionInfo(
                                name=method_name,
                                file_path=file_path,
                                start_line=start_line,
                                end_line=line_num,
                                size=size
                            ))
                            current_method = None
                        else:
                            current_method = (method_name, start_line, brace_count)
                    else:
                        # Look for new method declarations
                        match = JavaParser.METHOD_PATTERN.search(line)
                        if match:
                            method_name = match.group(1)
                            brace_count = line.count('{') - line.count('}')
                            
                            if brace_count == 0:
                                # Single-line method (rare but possible)
                                functions.append(FunctionInfo(
                                    name=method_name,
                                    file_path=file_path,
                                    start_line=line_num,
                                    end_line=line_num,
                                    size=1
                                ))
                            elif brace_count > 0:
                                # Multi-line method - start tracking
                                current_method = (method_name, line_num, brace_count)
                                
        except Exception as e:
            print(f"Warning: Could not read {file_path}: {e}")
        
        return functions


class CSharpParser:
    """Parser for C# methods."""
    
    # Compile regex pattern once for better performance
    # C# methods can have opening brace on same line or next line
    METHOD_PATTERN = re.compile(
        r'^\s*(?:public|private|protected|internal)?\s*(?:static)?\s*(?:virtual|override|abstract|sealed|async)?\s*'
        r'[\w<>\[\]?]+\s+(\w+)\s*\([^)]*\)\s*\{?\s*$'
    )
    
    @staticmethod
    def parse_functions(file_path: str) -> List[FunctionInfo]:
        """
        Parse C# file to extract methods.
        
        Uses a streaming approach to handle very large files efficiently
        without loading the entire file into memory.
        
        Supports methods with various modifiers including:
        - Access modifiers: public, private, protected, internal
        - Other modifiers: static, virtual, override, abstract, sealed, async
        - Generic return types and constraints
        
        Args:
            file_path: Path to the C# file
            
        Returns:
            List of FunctionInfo objects for all detected methods
        """
        functions = []
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                line_num = 0
                current_method = None  # (name, start_line, brace_count)
                pending_method = None  # (name, start_line) waiting for opening brace
                
                for line in f:  # Stream file line by line
                    line_num += 1
                    stripped = line.strip()
                    
                    # If we're currently tracking a method, update brace count
                    if current_method:
                        method_name, start_line, brace_count = current_method
                        brace_count += line.count('{') - line.count('}')
                        
                        if brace_count == 0:
                            # Method ended
                            size = line_num - start_line + 1
                            functions.append(FunctionInfo(
                                name=method_name,
                                file_path=file_path,
                                start_line=start_line,
                                end_line=line_num,
                                size=size
                            ))
                            current_method = None
                        else:
                            current_method = (method_name, start_line, brace_count)
                    
                    # Check if there's a pending method waiting for opening brace
                    elif pending_method and stripped == '{':
                        method_name, start_line = pending_method
                        # Start tracking the method
                        current_method = (method_name, start_line, 1)
                        pending_method = None
                    
                    # Look for new method declarations
                    else:
                        match = CSharpParser.METHOD_PATTERN.search(line)
                        if match:
                            method_name = match.group(1)
                            # Discard any pending method if we find a new declaration
                            pending_method = None
                            # Check if opening brace is on the same line
                            if '{' in line:
                                brace_count = line.count('{') - line.count('}')
                                if brace_count == 0:
                                    # Single-line method (rare)
                                    functions.append(FunctionInfo(
                                        name=method_name,
                                        file_path=file_path,
                                        start_line=line_num,
                                        end_line=line_num,
                                        size=1
                                    ))
                                else:
                                    # Multi-line method - start tracking
                                    current_method = (method_name, line_num, brace_count)
                            else:
                                # Waiting for opening brace on next line
                                pending_method = (method_name, line_num)
                                
        except Exception as e:
            print(f"Warning: Could not read {file_path}: {e}")
        
        return functions


def scan_single_repository(repo_path: str) -> Tuple[str, List[FunctionInfo]]:
    """
    Scan a single repository and return results.
    
    This function is designed to be called in parallel. It handles both remote
    repositories (which are cloned to a temporary directory) and local repositories.
    
    Args:
        repo_path: URL of a git repository or path to a local repository
        
    Returns:
        A tuple of (repository_name, list_of_functions). Returns (None, []) on error.
    """
    # Common directories to skip - using sets for O(1) lookup
    SKIP_DIRS = {
        'node_modules', '.git', 'target', 'build', 'out', 
        '__pycache__', 'venv', 'env', '.venv', 'site-packages',
        'dist', 'coverage', '.tox', '.pytest_cache', '.mypy_cache'
    }
    
    temp_dir = None
    try:
        # Clone or use local repo
        if repo_path.startswith('http://') or repo_path.startswith('https://') or repo_path.startswith('git@'):
            # It's a remote repository - clone it
            temp_dir = tempfile.mkdtemp(prefix='function_calculator_')
            
            print(f"Cloning repository: {repo_path}")
            try:
                subprocess.run(
                    ['git', 'clone', '--depth', '1', repo_path, temp_dir],
                    check=True,
                    capture_output=True,
                    text=True,
                    timeout=300  # 5 minute timeout for cloning
                )
                local_path = temp_dir
            except subprocess.TimeoutExpired:
                print(f"Error: Timeout cloning repository {repo_path} (exceeded 5 minutes)")
                return None, []
            except subprocess.CalledProcessError as e:
                print(f"Error cloning repository {repo_path}: {e}")
                return None, []
        else:
            # It's a local path
            if os.path.exists(repo_path):
                local_path = repo_path
            else:
                print(f"Error: Local path does not exist: {repo_path}")
                return None, []
        
        all_functions = []
        
        # Find all JavaScript/TypeScript files
        js_extensions = ['.js', '.jsx', '.ts', '.tsx', '.mjs']
        for ext in js_extensions:
            for file_path in Path(local_path).rglob(f'*{ext}'):
                # Skip common directories using set lookup
                if any(part in SKIP_DIRS for part in file_path.parts):
                    continue
                
                functions = JavaScriptParser.parse_functions(str(file_path))
                # Make paths relative to repo root
                for func in functions:
                    func.file_path = os.path.relpath(func.file_path, local_path)
                all_functions.extend(functions)
        
        # Find all Java files
        for file_path in Path(local_path).rglob('*.java'):
            # Skip common build directories using set lookup
            if any(part in SKIP_DIRS for part in file_path.parts):
                continue
            
            functions = JavaParser.parse_functions(str(file_path))
            # Make paths relative to repo root
            for func in functions:
                func.file_path = os.path.relpath(func.file_path, local_path)
            all_functions.extend(functions)
        
        # Find all Python files
        for file_path in Path(local_path).rglob('*.py'):
            # Skip common directories and virtual environments using set lookup
            if any(part in SKIP_DIRS for part in file_path.parts):
                continue
            
            functions = PythonParser.parse_functions(str(file_path))
            # Make paths relative to repo root
            for func in functions:
                func.file_path = os.path.relpath(func.file_path, local_path)
            all_functions.extend(functions)
        
        # Find all C# files
        for file_path in Path(local_path).rglob('*.cs'):
            # Skip common build directories using set lookup
            if any(part in SKIP_DIRS for part in file_path.parts):
                continue
            
            functions = CSharpParser.parse_functions(str(file_path))
            # Make paths relative to repo root
            for func in functions:
                func.file_path = os.path.relpath(func.file_path, local_path)
            all_functions.extend(functions)
        
        # Get repository name
        repo_name = os.path.basename(repo_path.rstrip('/').replace('.git', ''))
        if not repo_name:
            repo_name = 'repository'
        
        return repo_name, all_functions
    
    finally:
        # Cleanup temporary directory
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


class ExcelWriter:
    """Writes results to Excel (XLSX) file."""
    
    @staticmethod
    def write_results(repo_results: Dict[str, List[FunctionInfo]], output_file: str, 
                     top_n: int = 5, min_size: int = 1):
        """
        Write results to Excel (XLSX) file with each repo on a separate tab.
        
        Args:
            repo_results: Dictionary mapping repository names to lists of functions
            output_file: Path to the output Excel file
            top_n: Number of top functions to include per repository
            min_size: Minimum function size (in lines) to include
        """
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        for repo_name, functions in repo_results.items():
            # Filter by minimum size
            filtered_functions = [f for f in functions if f.size >= min_size]
            
            # Create sanitized sheet name (Excel has 31 char limit and special char restrictions)
            sheet_name = repo_name.replace('/', '_').replace('\\', '_')
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Add header row
            headers = ['Rank', 'Function Name', 'File Path', 'Start Line', 'End Line', 'Lines of Code']
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Sort functions by size (descending) and take top N
            top_functions = sorted(filtered_functions, key=lambda f: f.size, reverse=True)[:top_n]
            
            # Add data rows
            for rank, func in enumerate(top_functions, 1):
                ws.append([
                    rank,
                    func.name,
                    func.file_path,
                    func.start_line,
                    func.end_line,
                    func.size
                ])
            
            # Add summary statistics at the bottom
            # Calculate in a single pass for efficiency
            if filtered_functions:
                total = len(filtered_functions)
                total_size = 0
                smallest_size = float('inf')
                largest_size = 0
                
                for func in filtered_functions:
                    total_size += func.size
                    if func.size < smallest_size:
                        smallest_size = func.size
                    if func.size > largest_size:
                        largest_size = func.size
                
                ws.append([])  # Empty row
                ws.append(['Summary Statistics'])
                ws.append(['Total Functions Found:', total])
                ws.append(['Average Function Size:', f"{total_size / total:.1f} lines"])
                ws.append(['Largest Function:', largest_size])
                ws.append(['Smallest Function:', smallest_size])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
        
        wb.save(output_file)
        print(f"\nResults saved to: {output_file}")


class JSONWriter:
    """Writes results to JSON file."""
    
    @staticmethod
    def write_results(repo_results: Dict[str, List[FunctionInfo]], output_file: str,
                     top_n: int = 5, min_size: int = 1):
        """
        Write results to JSON file.
        
        Args:
            repo_results: Dictionary mapping repository names to lists of functions
            output_file: Path to the output JSON file
            top_n: Number of top functions to include per repository
            min_size: Minimum function size (in lines) to include
        """
        output_data = {}
        
        for repo_name, functions in repo_results.items():
            # Filter by minimum size
            filtered_functions = [f for f in functions if f.size >= min_size]
            
            # Sort functions by size (descending) and take top N
            top_functions = sorted(filtered_functions, key=lambda f: f.size, reverse=True)[:top_n]
            
            # Calculate summary statistics in a single pass
            summary = {}
            if filtered_functions:
                total = len(filtered_functions)
                total_size = 0
                smallest_size = float('inf')
                largest_size = 0
                
                for func in filtered_functions:
                    total_size += func.size
                    if func.size < smallest_size:
                        smallest_size = func.size
                    if func.size > largest_size:
                        largest_size = func.size
                
                summary = {
                    'total_functions': total,
                    'average_size': round(total_size / total, 1),
                    'largest_function_size': largest_size,
                    'smallest_function_size': smallest_size
                }
            
            output_data[repo_name] = {
                'summary': summary,
                'top_functions': [f.to_dict() for f in top_functions]
            }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        print(f"\nResults saved to: {output_file}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Find the largest functions in git repositories (Node.js, Java, and Python)'
    )
    parser.add_argument(
        'repositories',
        nargs='*',
        help='Git repository URLs or local paths to scan'
    )
    parser.add_argument(
        '-i', '--input-file',
        help='File containing list of repository URLs/paths (one per line)'
    )
    parser.add_argument(
        '-o', '--output',
        default='function_sizes.xlsx',
        help='Output file name (default: function_sizes.xlsx). Use .json extension for JSON format.'
    )
    parser.add_argument(
        '-f', '--format',
        choices=['xlsx', 'json', 'auto'],
        default='auto',
        help='Output format (default: auto - detect from file extension)'
    )
    parser.add_argument(
        '-j', '--jobs',
        type=int,
        default=4,
        help='Number of parallel jobs (default: 4)'
    )
    parser.add_argument(
        '-n', '--top-n',
        type=int,
        default=5,
        help='Number of top largest functions to report per repository (default: 5)'
    )
    parser.add_argument(
        '-m', '--min-size',
        type=int,
        default=1,
        help='Minimum function size in lines to include (default: 1)'
    )
    
    args = parser.parse_args()
    
    # Validate jobs parameter
    if args.jobs < 1:
        print("Error: Number of parallel jobs must be at least 1")
        sys.exit(1)
    
    # Validate top-n parameter
    if args.top_n < 1:
        print("Error: Number of top functions must be at least 1")
        sys.exit(1)
    
    # Validate min-size parameter
    if args.min_size < 1:
        print("Error: Minimum function size must be at least 1")
        sys.exit(1)
    
    # Determine output format
    output_format = args.format
    if output_format == 'auto':
        if args.output.endswith('.json'):
            output_format = 'json'
        elif args.output.endswith('.xlsx'):
            output_format = 'xlsx'
        else:
            print("Error: Output file must have .xlsx or .json extension")
            sys.exit(1)
    
    # Validate output file extension matches format
    if output_format == 'xlsx' and not args.output.endswith('.xlsx'):
        print("Error: Output file must have .xlsx extension when using xlsx format")
        sys.exit(1)
    elif output_format == 'json' and not args.output.endswith('.json'):
        print("Error: Output file must have .json extension when using json format")
        sys.exit(1)
    
    # Collect repositories from command line and/or input file
    repositories = list(args.repositories) if args.repositories else []
    
    if args.input_file:
        try:
            with open(args.input_file, 'r') as f:
                for line in f:
                    line = line.strip()
                    # Skip empty lines and comments
                    if line and not line.startswith('#'):
                        repositories.append(line)
        except FileNotFoundError:
            print(f"Error: Input file not found: {args.input_file}")
            sys.exit(1)
        except Exception as e:
            print(f"Error reading input file: {e}")
            sys.exit(1)
    
    if not repositories:
        parser.print_help()
        print("\nError: No repositories specified. Provide repositories via command line or --input-file")
        sys.exit(1)
    
    print(f"Scanning {len(repositories)} repositories using {args.jobs} parallel jobs...")
    print(f"Configuration: Top {args.top_n} functions, Minimum size: {args.min_size} lines")
    print(f"{'='*60}\n")
    
    repo_results = {}
    completed_count = 0
    total_repos = len(repositories)
    
    # Show initial progress bar
    print_progress_bar(0, total_repos, prefix='Progress:', suffix='Complete')
    
    # Process repositories in parallel
    with ProcessPoolExecutor(max_workers=args.jobs) as executor:
        # Submit all tasks
        future_to_repo = {executor.submit(scan_single_repository, repo): repo for repo in repositories}
        
        # Process completed tasks
        for future in as_completed(future_to_repo):
            repo = future_to_repo[future]
            try:
                repo_name, functions = future.result()
                
                if repo_name is not None:
                    repo_results[repo_name] = functions
                    completed_count += 1
                    
                    # Update progress bar
                    print_progress_bar(completed_count, total_repos, 
                                     prefix='Progress:', 
                                     suffix=f'Complete ({completed_count}/{total_repos})')
                    
                    # Filter by minimum size for display
                    filtered = [f for f in functions if f.size >= args.min_size]
                    
                    # Print summary for this repository
                    top_n_display = sorted(filtered, key=lambda f: f.size, reverse=True)[:args.top_n]
                    print(f"\n✓ Repository: {repo}")
                    print(f"  Found {len(functions)} functions ({len(filtered)} >= {args.min_size} lines). Top {args.top_n} largest:")
                    for i, func in enumerate(top_n_display, 1):
                        print(f"    {i}. {func.name} ({func.size} lines) - {func.file_path}")
                    
                    # Print progress bar again after summary (if not last repo)
                    if completed_count < total_repos:
                        print()  # Empty line before progress bar
                        print_progress_bar(completed_count, total_repos, 
                                         prefix='Progress:', 
                                         suffix=f'Complete ({completed_count}/{total_repos})')
            except Exception as e:
                completed_count += 1
                print_progress_bar(completed_count, total_repos, 
                                 prefix='Progress:', 
                                 suffix=f'Complete ({completed_count}/{total_repos})')
                print(f"\n✗ Error processing repository {repo}: {e}")
                if completed_count < total_repos:
                    print()
                    print_progress_bar(completed_count, total_repos, 
                                     prefix='Progress:', 
                                     suffix=f'Complete ({completed_count}/{total_repos})')
    
    
    # Write results to file
    print()  # Add blank line after progress bar
    if repo_results:
        if output_format == 'json':
            JSONWriter.write_results(repo_results, args.output, args.top_n, args.min_size)
        else:
            ExcelWriter.write_results(repo_results, args.output, args.top_n, args.min_size)
        print(f"\n{'='*60}")
        print(f"✓ Done! Check {args.output} for detailed results.")
        print(f"{'='*60}")
    else:
        print("\n✗ No results to write. Please check the repository paths/URLs.")


if __name__ == '__main__':
    main()
